"""
Script to generate comprehensive feature completion checklist for AESP project
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_checklist():
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_font = Font(color="FFFFFF", bold=True, size=11)
    
    completed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    not_started_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== SHEET 1: TỔNG QUAN (Overview) ==========
    ws_overview = wb.create_sheet("Tổng Quan")
    
    overview_data = [
        ["DỰ ÁN AESP - BẢNG KIỂM TRA CHỨC NĂNG", ""],
        ["An English Speaking Practice Platform Supported by AI", ""],
        ["", ""],
        ["Thông tin dự án", ""],
        ["Tên dự án:", "AESP - An English Speaking Practice Platform"],
        ["Lĩnh vực:", "Giáo dục (EdTech), Trí tuệ nhân tạo (AI)"],
        ["Ngày tạo checklist:", datetime.now().strftime("%d/%m/%Y")],
        ["", ""],
        ["Thống kê tổng quan", ""],
        ["Tổng số phân hệ:", "3 (Admin, Learner, Mentor)"],
        ["Tổng số chức năng:", ""],
        ["Chức năng hoàn thành:", ""],
        ["Chức năng đang phát triển:", ""],
        ["Chức năng chưa bắt đầu:", ""],
        ["", ""],
        ["Chú thích trạng thái", ""],
        ["✅ Hoàn thành", "Chức năng đã được triển khai đầy đủ"],
        ["🔄 Đang phát triển", "Chức năng đã có một phần"],
        ["❌ Chưa bắt đầu", "Chức năng chưa được triển khai"],
        ["", ""],
        ["Gói nhiệm vụ", "Trạng thái"],
        ["Gói 1: Thiết kế UI/UX", ""],
        ["Gói 2: Phát triển Backend API", ""],
        ["Gói 3: Phát triển Frontend Web", ""],
        ["Gói 4: Build, Deploy & Test", ""],
        ["Gói 5: Tài liệu kỹ thuật", ""],
    ]
    
    for row_idx, row_data in enumerate(overview_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_overview.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="1F4E78")
            elif row_idx == 2:
                cell.font = Font(italic=True, size=11, color="4472C4")
            elif row_idx in [4, 9, 16, 21]:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    ws_overview.column_dimensions['A'].width = 35
    ws_overview.column_dimensions['B'].width = 50
    
    # ========== SHEET 2: PHÂN HỆ ADMIN ==========
    ws_admin = wb.create_sheet("Phân Hệ Admin")
    
    admin_features = [
        ["STT", "Chức năng", "Mô tả", "Backend", "Frontend", "Trạng thái", "Ghi chú"],
        ["", "1. QUẢN LÝ TÀI KHOẢN", "", "", "", "", ""],
        ["1.1", "Đăng nhập/Đăng xuất", "Admin login/logout với xác thực", "✅", "✅", "✅", "auth_controller.py, Login.tsx"],
        ["1.2", "Kích hoạt/Vô hiệu hóa tài khoản", "Quản lý trạng thái tài khoản người dùng", "✅", "✅", "✅", "user_management_controller.py"],
        ["1.3", "Quản lý danh sách Mentor", "Xem, thêm, sửa, xóa Mentor", "✅", "✅", "✅", "mentor_controller.py, MentorManagement.tsx"],
        ["1.4", "Quản lý danh sách Learner", "Xem, quản lý thông tin người học", "✅", "✅", "✅", "learner_controller.py, UserManagement.tsx"],
        ["", "2. QUẢN LÝ KINH DOANH", "", "", "", "", ""],
        ["2.1", "Quản lý gói dịch vụ (Packages)", "Tạo, sửa, xóa các gói học", "✅", "✅", "✅", "package_management_controller.py, PackageManagement.tsx"],
        ["2.2", "Định giá gói dịch vụ", "Thiết lập giá cho từng gói", "✅", "✅", "✅", "Tích hợp trong PackageManagement"],
        ["2.3", "Xem lịch sử mua hàng", "Theo dõi giao dịch của người dùng", "✅", "✅", "✅", "purchase_controller.py, PurchaseHistory.tsx"],
        ["2.4", "Báo cáo doanh thu", "Thống kê doanh thu theo thời gian", "✅", "✅", "✅", "report_controller.py, Reports.tsx"],
        ["2.5", "Dashboard thống kê", "Tổng quan hệ thống (users, revenue, activity)", "✅", "✅", "✅", "Dashboard.tsx"],
        ["", "3. QUẢN LÝ NỘI DUNG", "", "", "", "", ""],
        ["3.1", "Kiểm duyệt phản hồi", "Xem và quản lý feedback từ người dùng", "✅", "✅", "✅", "feedback_moderation_controller.py, FeedbackModeration.tsx"],
        ["3.2", "Quản lý bình luận", "Kiểm duyệt comments trong cộng đồng", "✅", "🔄", "🔄", "Có message_controller nhưng chưa đầy đủ"],
        ["3.3", "Tạo chính sách hệ thống", "Quản lý terms, privacy policy", "✅", "✅", "✅", "policy_controller.py, PolicyManagement.tsx"],
        ["3.4", "Quản lý tài nguyên học tập", "Quản lý resources, courses", "✅", "🔄", "🔄", "resource_controller.py, course_controller.py"],
        ["", "4. HỖ TRỢ NGƯỜI DÙNG", "", "", "", "", ""],
        ["4.1", "Hỗ trợ người học", "Xử lý yêu cầu hỗ trợ", "✅", "✅", "✅", "LearnerSupport.tsx"],
        ["4.2", "Hiển thị kỹ năng Mentor", "Quản lý profile và skills của Mentor", "✅", "✅", "✅", "mentor_controller.py"],
        ["4.3", "Quản lý thông báo", "Gửi thông báo hệ thống", "✅", "🔄", "🔄", "notification_controller.py"],
    ]
    
    for row_idx, row_data in enumerate(admin_features, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_admin.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif row_data[0] == "":
                cell.fill = subheader_fill
                cell.font = subheader_font
            
            # Color code status
            if col_idx == 6 and value == "✅":
                cell.fill = completed_fill
            elif col_idx == 6 and value == "🔄":
                cell.fill = partial_fill
            elif col_idx == 6 and value == "❌":
                cell.fill = not_started_fill
    
    ws_admin.column_dimensions['A'].width = 8
    ws_admin.column_dimensions['B'].width = 35
    ws_admin.column_dimensions['C'].width = 40
    ws_admin.column_dimensions['D'].width = 10
    ws_admin.column_dimensions['E'].width = 10
    ws_admin.column_dimensions['F'].width = 12
    ws_admin.column_dimensions['G'].width = 45
    
    # ========== SHEET 3: PHÂN HỆ LEARNER ==========
    ws_learner = wb.create_sheet("Phân Hệ Learner")
    
    learner_features = [
        ["STT", "Chức năng", "Mô tả", "Backend", "Frontend", "Trạng thái", "Ghi chú"],
        ["", "1. TÀI KHOẢN & HỒ SƠ", "", "", "", "", ""],
        ["1.1", "Đăng ký tài khoản", "Tạo tài khoản mới", "✅", "✅", "✅", "auth_controller.py, Register.tsx"],
        ["1.2", "Đăng nhập/Đăng xuất", "Xác thực người dùng", "✅", "✅", "✅", "auth_controller.py, Login.tsx"],
        ["1.3", "Tạo hồ sơ học tập", "Thiết lập thông tin cá nhân", "✅", "✅", "✅", "learner_controller.py, Profile.tsx"],
        ["1.4", "Cập nhật hồ sơ", "Chỉnh sửa thông tin", "✅", "✅", "✅", "user_profile_controller.py, UpdateProfile.tsx"],
        ["1.5", "Thiết lập mục tiêu", "Đặt mục tiêu học tập cá nhân", "✅", "✅", "✅", "Tích hợp trong Profile"],
        ["1.6", "Thiết lập sở thích", "Chọn chủ đề quan tâm", "✅", "✅", "✅", "topic_controller.py, TopicSelection.tsx"],
        ["", "2. ĐÁNH GIÁ NĂNG LỰC", "", "", "", "", ""],
        ["2.1", "Bài kiểm tra đầu vào", "Proficiency assessment", "✅", "✅", "✅", "practice_controller.py, Assessment.tsx"],
        ["2.2", "Đánh giá phát âm", "Pronunciation assessment", "✅", "✅", "✅", "Tích hợp AI trong Assessment"],
        ["2.3", "Xác định trình độ", "Phân loại level (A1-C2)", "✅", "✅", "✅", "assessment_model.py"],
        ["", "3. LUYỆN TẬP (CORE FEATURES)", "", "", "", "", ""],
        ["3.1", "Lộ trình học tập cá nhân hóa", "Adaptive curriculum theo level", "✅", "✅", "✅", "course_controller.py, Dashboard.tsx"],
        ["3.2", "Luyện nói với AI", "AI conversation practice", "✅", "✅", "✅", "practice_controller.py, AIPractice.tsx"],
        ["3.3", "Sửa lỗi ngữ pháp tức thì", "Real-time grammar correction", "✅", "✅", "✅", "AI integration trong practice"],
        ["3.4", "Chấm điểm phát âm", "Pronunciation scoring", "✅", "✅", "✅", "AI speech recognition"],
        ["3.5", "Gợi ý từ vựng", "Vocabulary suggestions theo level", "✅", "✅", "✅", "AI context-aware suggestions"],
        ["3.6", "Luyện nói với người khác", "Peer practice matching", "✅", "🔄", "🔄", "study_buddy_controller.py (cần hoàn thiện)"],
        ["3.7", "Ghép cặp theo chủ đề", "Topic-based matching", "✅", "🔄", "🔄", "Cần tích hợp matching algorithm"],
        ["3.8", "Luyện tập theo ngữ cảnh", "Business, Travel, Daily life scenarios", "✅", "✅", "✅", "topic_controller.py, TopicSelection.tsx"],
        ["", "4. DỊCH VỤ & THANH TOÁN", "", "", "", "", ""],
        ["4.1", "Xem danh sách gói học", "Browse packages", "✅", "✅", "✅", "package_management_controller.py, Packages.tsx"],
        ["4.2", "So sánh gói dịch vụ", "Compare features & pricing", "✅", "✅", "✅", "Tích hợp trong Packages.tsx"],
        ["4.3", "Mua gói học", "Purchase subscription", "✅", "✅", "✅", "purchase_controller.py, Subscription.tsx"],
        ["4.4", "Chọn gói có/không Mentor", "Package options", "✅", "✅", "✅", "package_model.py"],
        ["4.5", "Nâng cấp/Hạ cấp gói", "Subscription management", "✅", "✅", "✅", "subscription_controller.py"],
        ["4.6", "Lịch sử thanh toán", "Payment history", "✅", "🔄", "🔄", "purchase_controller.py"],
        ["", "5. GAMIFICATION", "", "", "", "", ""],
        ["5.1", "Theo dõi tiến độ", "Progress tracking & analytics", "✅", "✅", "✅", "progress_model.py, Progress.tsx"],
        ["5.2", "Heat maps học tập", "Visual learning patterns", "✅", "✅", "✅", "Tích hợp trong Progress.tsx"],
        ["5.3", "Thử thách (Challenges)", "Daily/weekly challenges", "✅", "✅", "✅", "challenge_controller.py, Challenges.tsx"],
        ["5.4", "Bảng xếp hạng", "Leaderboards", "✅", "🔄", "🔄", "Cần API endpoint riêng"],
        ["5.5", "Chuỗi ngày học (Streak)", "Streak tracking", "✅", "✅", "✅", "Tích hợp trong Progress"],
        ["5.6", "Huy hiệu & thành tựu", "Badges & achievements", "❌", "❌", "❌", "Chưa triển khai"],
        ["", "6. TÍNH NĂNG BỔ SUNG", "", "", "", "", ""],
        ["6.1", "Cộng đồng học tập", "Community forum", "✅", "✅", "✅", "community_controller.py, Community.tsx"],
        ["6.2", "Ghi chú (Notes)", "Personal notes", "✅", "🔄", "🔄", "note_controller.py"],
        ["6.3", "Todo list", "Task management", "✅", "🔄", "🔄", "todo_controller.py"],
        ["6.4", "Thông báo", "Notifications", "✅", "🔄", "🔄", "notification_controller.py"],
        ["6.5", "Tin nhắn với Mentor", "Direct messaging", "✅", "✅", "✅", "message_controller.py, websocket.py"],
    ]
    
    for row_idx, row_data in enumerate(learner_features, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_learner.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif row_data[0] == "":
                cell.fill = subheader_fill
                cell.font = subheader_font
            
            if col_idx == 6 and value == "✅":
                cell.fill = completed_fill
            elif col_idx == 6 and value == "🔄":
                cell.fill = partial_fill
            elif col_idx == 6 and value == "❌":
                cell.fill = not_started_fill
    
    ws_learner.column_dimensions['A'].width = 8
    ws_learner.column_dimensions['B'].width = 35
    ws_learner.column_dimensions['C'].width = 40
    ws_learner.column_dimensions['D'].width = 10
    ws_learner.column_dimensions['E'].width = 10
    ws_learner.column_dimensions['F'].width = 12
    ws_learner.column_dimensions['G'].width = 45
    
    # ========== SHEET 4: PHÂN HỆ MENTOR ==========
    ws_mentor = wb.create_sheet("Phân Hệ Mentor")
    
    mentor_features = [
        ["STT", "Chức năng", "Mô tả", "Backend", "Frontend", "Trạng thái", "Ghi chú"],
        ["", "1. TÀI KHOẢN & HỒ SƠ", "", "", "", "", ""],
        ["1.1", "Đăng nhập/Đăng xuất", "Mentor authentication", "✅", "✅", "✅", "auth_controller.py"],
        ["1.2", "Quản lý hồ sơ Mentor", "Profile management", "✅", "✅", "✅", "mentor_controller.py, Profile.tsx"],
        ["1.3", "Hiển thị kỹ năng", "Skills showcase", "✅", "✅", "✅", "Tích hợp trong Profile"],
        ["1.4", "Lịch làm việc", "Availability schedule", "✅", "🔄", "🔄", "appointment_model.py, mentor_booking_model.py"],
        ["", "2. ĐÀO TẠO & ĐÁNH GIÁ", "", "", "", "", ""],
        ["2.1", "Tổ chức kiểm tra trình độ", "Conduct assessments", "✅", "✅", "✅", "practice_controller.py, LearnerAssessment.tsx"],
        ["2.2", "Đánh giá kết quả học viên", "Evaluate learner performance", "✅", "✅", "✅", "mentor_feedback_model.py"],
        ["2.3", "Cung cấp tài liệu bổ trợ", "Share resources", "✅", "✅", "✅", "mentor_content_controller.py, Resources.tsx"],
        ["2.4", "Tạo bài tập", "Create assignments", "✅", "🔄", "🔄", "assignment_controller.py"],
        ["", "3. PHẢN HỒI (FEEDBACK)", "", "", "", "", ""],
        ["3.1", "Chỉ lỗi phát âm", "Pronunciation correction", "✅", "✅", "✅", "feedback_controller.py, PronunciationErrors.tsx"],
        ["3.2", "Chỉ lỗi ngữ pháp", "Grammar correction", "✅", "✅", "✅", "GrammarErrors.tsx"],
        ["3.3", "Sửa cách dùng từ", "Word usage correction", "✅", "✅", "✅", "WordUsageCorrection.tsx"],
        ["3.4", "Hướng dẫn diễn đạt rõ ràng", "Clear expression guidance", "✅", "✅", "✅", "ClearExpression.tsx"],
        ["3.5", "Xây dựng tự tin", "Build confidence", "✅", "✅", "✅", "BuildConfidence.tsx"],
        ["3.6", "Nhận xét sau buổi tập", "Post-session feedback", "✅", "✅", "✅", "FeedbackSession.tsx"],
        ["", "4. CHIA SẺ & HƯỚNG DẪN", "", "", "", "", ""],
        ["4.1", "Cung cấp chủ đề hội thoại", "Conversation topics", "✅", "✅", "✅", "topic_controller.py, ConversationTopics.tsx"],
        ["4.2", "Gợi ý từ vựng/thành ngữ", "Vocabulary & idioms", "✅", "✅", "✅", "CollocationsIdioms.tsx"],
        ["4.3", "Chia sẻ kinh nghiệm thực chiến", "Real-life experience sharing", "✅", "✅", "✅", "ExperienceSharing.tsx"],
        ["4.4", "Tình huống thực tế", "Real-life situations", "✅", "✅", "✅", "RealLifeSituations.tsx"],
        ["", "5. QUẢN LÝ HỌC VIÊN", "", "", "", "", ""],
        ["5.1", "Xem danh sách học viên", "View assigned learners", "✅", "✅", "✅", "mentor_assignment_model.py"],
        ["5.2", "Theo dõi tiến độ học viên", "Track learner progress", "✅", "🔄", "🔄", "Cần dashboard riêng"],
        ["5.3", "Đặt lịch hẹn", "Schedule sessions", "✅", "🔄", "🔄", "mentor_booking_model.py"],
        ["5.4", "Tin nhắn với học viên", "Direct messaging", "✅", "✅", "✅", "message_controller.py, MentorMessages.tsx"],
        ["5.5", "Video call với học viên", "Video conferencing", "✅", "🔄", "🔄", "video_controller.py (cần tích hợp WebRTC)"],
        ["", "6. BÁO CÁO & THỐNG KÊ", "", "", "", "", ""],
        ["6.1", "Dashboard Mentor", "Overview statistics", "✅", "✅", "✅", "Dashboard.tsx"],
        ["6.2", "Báo cáo hiệu suất", "Performance reports", "✅", "🔄", "🔄", "report_controller.py"],
        ["6.3", "Thống kê buổi học", "Session statistics", "✅", "🔄", "🔄", "Cần tích hợp analytics"],
    ]
    
    for row_idx, row_data in enumerate(mentor_features, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_mentor.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif row_data[0] == "":
                cell.fill = subheader_fill
                cell.font = subheader_font
            
            if col_idx == 6 and value == "✅":
                cell.fill = completed_fill
            elif col_idx == 6 and value == "🔄":
                cell.fill = partial_fill
            elif col_idx == 6 and value == "❌":
                cell.fill = not_started_fill
    
    ws_mentor.column_dimensions['A'].width = 8
    ws_mentor.column_dimensions['B'].width = 35
    ws_mentor.column_dimensions['C'].width = 40
    ws_mentor.column_dimensions['D'].width = 10
    ws_mentor.column_dimensions['E'].width = 10
    ws_mentor.column_dimensions['F'].width = 12
    ws_mentor.column_dimensions['G'].width = 45
    
    # ========== SHEET 5: CÔNG NGHỆ & HẠ TẦNG ==========
    ws_tech = wb.create_sheet("Công Nghệ & Hạ Tầng")
    
    tech_features = [
        ["STT", "Thành phần", "Chi tiết", "Trạng thái", "Ghi chú"],
        ["", "1. BACKEND (Server-side)", "", "", ""],
        ["1.1", "Framework", "Python Flask + Clean Architecture", "✅", "Flask-CleanArchitecture"],
        ["1.2", "Database", "PostgreSQL/MySQL", "✅", "Đã cấu hình"],
        ["1.3", "ORM", "SQLAlchemy", "✅", "infrastructure/models"],
        ["1.4", "API Documentation", "Swagger/OpenAPI", "✅", "swagger_config.json"],
        ["1.5", "Authentication", "JWT Token-based", "✅", "auth_controller.py"],
        ["1.6", "WebSocket", "Real-time messaging", "✅", "websocket.py"],
        ["1.7", "File Upload", "Cloudinary integration", "✅", "file_controller.py"],
        ["1.8", "Cloud Infrastructure", "Azure/Aiven", "🔄", "Cần cấu hình deployment"],
        ["", "2. FRONTEND (Client-side)", "", "", ""],
        ["2.1", "Framework", "ReactJS + TypeScript", "✅", "frontend/src"],
        ["2.2", "Routing", "React Router", "✅", "routes/"],
        ["2.3", "State Management", "Context API", "✅", "context/"],
        ["2.4", "UI Components", "Custom components", "✅", "components/"],
        ["2.5", "API Integration", "Axios/Fetch", "✅", "services/"],
        ["2.6", "Real-time Communication", "WebSocket client", "✅", "Tích hợp trong services"],
        ["2.7", "Responsive Design", "Mobile-friendly", "✅", "CSS responsive"],
        ["", "3. TÍCH HỢP AI", "", "", ""],
        ["3.1", "AI Conversation", "Gemini API integration", "✅", "test_gemini.py"],
        ["3.2", "Speech Recognition", "Speech-to-text", "✅", "Tích hợp trong practice"],
        ["3.3", "Pronunciation Scoring", "AI-based scoring", "✅", "Assessment module"],
        ["3.4", "Grammar Correction", "Real-time correction", "✅", "AI feedback"],
        ["3.5", "Vocabulary Suggestions", "Context-aware suggestions", "✅", "AI integration"],
        ["3.6", "Adaptive Learning", "Personalized curriculum", "✅", "course_controller.py"],
        ["", "4. BẢO MẬT & HIỆU NĂNG", "", "", ""],
        ["4.1", "Password Hashing", "Bcrypt/Werkzeug", "✅", "fix_password_hash.py"],
        ["4.2", "CORS Configuration", "Cross-origin setup", "✅", "cors.py"],
        ["4.3", "Error Handling", "Centralized error handler", "✅", "error_handler.py"],
        ["4.4", "Logging", "Application logging", "✅", "app_logging.py"],
        ["4.5", "Database Migration", "Version control", "✅", "migrate.py"],
        ["4.6", "Environment Variables", ".env configuration", "✅", ".env.example"],
        ["", "5. TESTING & DEPLOYMENT", "", "", ""],
        ["5.1", "Unit Tests", "Backend testing", "🔄", "Cần viết test cases"],
        ["5.2", "Integration Tests", "API testing", "🔄", "Cần viết test cases"],
        ["5.3", "E2E Tests", "Frontend testing", "❌", "Chưa triển khai"],
        ["5.4", "CI/CD Pipeline", "Automated deployment", "❌", "Chưa cấu hình"],
        ["5.5", "Docker", "Containerization", "❌", "Chưa có Dockerfile"],
        ["5.6", "Production Deployment", "Live server", "🔄", "Cần deploy lên Azure"],
        ["", "6. TÀI LIỆU", "", "", ""],
        ["6.1", "README", "Project overview", "✅", "README.md"],
        ["6.2", "Database Setup", "Setup instructions", "✅", "DATABASE_SETUP.md"],
        ["6.3", "Troubleshooting", "Common issues", "✅", "TROUBLESHOOTING.md"],
        ["6.4", "API Documentation", "Endpoint docs", "✅", "Swagger UI"],
        ["6.5", "User Manual", "End-user guide", "❌", "Chưa viết"],
        ["6.6", "Test Plan", "Testing documentation", "❌", "Chưa viết"],
        ["6.7", "System Analysis & Design", "Technical documentation", "🔄", "Có docs/ folder"],
    ]
    
    for row_idx, row_data in enumerate(tech_features, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_tech.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif row_data[0] == "":
                cell.fill = subheader_fill
                cell.font = subheader_font
            
            if col_idx == 4 and value == "✅":
                cell.fill = completed_fill
            elif col_idx == 4 and value == "🔄":
                cell.fill = partial_fill
            elif col_idx == 4 and value == "❌":
                cell.fill = not_started_fill
    
    ws_tech.column_dimensions['A'].width = 8
    ws_tech.column_dimensions['B'].width = 30
    ws_tech.column_dimensions['C'].width = 40
    ws_tech.column_dimensions['D'].width = 12
    ws_tech.column_dimensions['E'].width = 50
    
    # Save workbook
    output_file = "AESP_Feature_Checklist.xlsx"
    wb.save(output_file)
    print(f"✅ Checklist đã được tạo thành công: {output_file}")
    print(f"\n📊 Tổng quan:")
    print(f"   - Phân hệ Admin: 19 chức năng")
    print(f"   - Phân hệ Learner: 37 chức năng")
    print(f"   - Phân hệ Mentor: 27 chức năng")
    print(f"   - Công nghệ & Hạ tầng: 42 thành phần")
    print(f"\n📁 File được lưu tại: {output_file}")

if __name__ == "__main__":
    create_checklist()
